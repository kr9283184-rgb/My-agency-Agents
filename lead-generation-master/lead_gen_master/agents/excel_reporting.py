import os
from typing import Optional
from datetime import datetime
from lead_gen_master.agents.base_agent import BaseAgent

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


class ExcelReportingAgent(BaseAgent):
    def __init__(self, memory=None):
        super().__init__(memory)

    def generate_report(
        self,
        leads: list[dict],
        audits: Optional[list[dict]] = None,
        output_path: Optional[str] = None,
    ) -> str:
        self.log("Generating Excel report")

        if output_path is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                self.memory.output_dir,
                f"lead_report_{ts}.xlsx",
            )

        qualified = [l for l in leads if l.get("lead_score", 0) >= 50]
        high = [l for l in leads if l.get("priority") == "High"]

        if pd is not None:
            self._with_pandas(
                leads, qualified, high,
                audits, output_path,
            )
        elif Workbook is not None:
            self._with_openpyxl(
                leads, qualified, high,
                audits, output_path,
            )
        else:
            csv_path = output_path.replace(".xlsx", ".csv")
            self._write_csv(leads, csv_path)
            return csv_path

        self.log(f"Report saved: {output_path}")
        return output_path

    def _with_pandas(
        self,
        leads: list[dict],
        qualified: list[dict],
        high: list[dict],
        audits: Optional[list[dict]],
        output_path: str,
    ):
        with pd.ExcelWriter(
            output_path, engine="openpyxl"
        ) as writer:
            df_all = pd.DataFrame(leads)
            df_all.to_excel(
                writer, sheet_name="All Leads", index=False
            )

            df_qual = pd.DataFrame(qualified)
            df_qual.to_excel(
                writer, sheet_name="Qualified Leads", index=False
            )

            df_high = pd.DataFrame(high)
            df_high.to_excel(
                writer, sheet_name="High Priority Leads", index=False
            )

            if audits:
                df_audit = pd.DataFrame(audits)
            else:
                df_audit = pd.DataFrame(
                    [{"info": "No audits performed"}]
                )
            df_audit.to_excel(
                writer, sheet_name="Website Audit Results", index=False
            )

            summary = pd.DataFrame([
                {
                    "Metric": "Total Leads Found",
                    "Value": len(leads),
                },
                {
                    "Metric": "Total Qualified Leads",
                    "Value": len(qualified),
                },
                {
                    "Metric": "High Priority Leads",
                    "Value": len(high),
                },
                {
                    "Metric": "Top Industry",
                    "Value": self._top_industry(leads),
                },
                {
                    "Metric": "Generated At",
                    "Value": datetime.now().isoformat(),
                },
            ])
            summary.to_excel(
                writer, sheet_name="Daily Summary", index=False
            )

    def _with_openpyxl(
        self,
        leads: list[dict],
        qualified: list[dict],
        high: list[dict],
        audits: Optional[list[dict]],
        output_path: str,
    ):
        wb = Workbook()
        self._write_sheet(
            wb.active, "All Leads", leads
        )
        self._write_sheet(
            wb.create_sheet(), "Qualified Leads", qualified
        )
        self._write_sheet(
            wb.create_sheet(), "High Priority Leads", high
        )
        self._write_sheet(
            wb.create_sheet(), "Website Audit Results",
            audits or [{"info": "No audits performed"}],
        )
        self._write_sheet(
            wb.create_sheet(), "Daily Summary",
            [
                {"Metric": "Total Leads Found", "Value": len(leads)},
                {"Metric": "Qualified Leads", "Value": len(qualified)},
                {"Metric": "High Priority", "Value": len(high)},
                {"Metric": "Top Industry", "Value": self._top_industry(leads)},
                {"Metric": "Date", "Value": datetime.now().isoformat()},
            ],
        )
        wb.save(output_path)

    def _write_sheet(self, ws, title: str, data: list[dict]):
        ws.title = title
        if not data:
            ws.cell(row=1, column=1, value="No data")
            return

        headers = list(data[0].keys())
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        header_font = Font(
            color="FFFFFF", bold=True
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(
                row=1, column=col_idx, value=header
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=record.get(header, ""),
                )

        ws.auto_filter.ref = ws.dimensions

    def _write_csv(
        self, leads: list[dict], output_path: str
    ):
        import csv

        if not leads:
            with open(output_path, "w") as f:
                f.write("No leads found\n")
            return

        headers = list(leads[0].keys())
        with open(output_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(leads)

    def _top_industry(self, leads: list[dict]) -> str:
        from collections import Counter

        industries = [
            l.get("industry", "Unknown")
            for l in leads
            if l.get("industry")
        ]
        if not industries:
            return "N/A"
        return Counter(industries).most_common(1)[0][0]
